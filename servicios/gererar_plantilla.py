from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

def crear_plantilla_excel():
    os.makedirs(os.path.join('static', 'assets'), exist_ok=True)
    ruta_guardado = os.path.join('static', 'assets', 'plantilla.xlsx')
    
    # Crear el libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Definir los encabezados de las columnas
    encabezados = ['fecha', 'descripcion', 'categoria', 'monto', 'tipo']
    ws.append(encabezados)
    gris_relleno = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    negrita = Font(bold=True)
    for celda in ws[1]:
        celda.fill = gris_relleno
        celda.font = negrita
    #Ejemplo:
    ws.append(['2026-07-01', 'Ejemplo ingreso', '', 1500000, 'ingreso'])
    ws.append(['2026-07-02', 'Mercado quincena', 'Alimentación', 120000, 'gasto'])

    ws.column_dimensions['B'].width = 25 # Descripcion mas ancha
    ws.column_dimensions['C'].width = 15 # Categoría
    
    #Guardar el archivo
    wb.save(ruta_guardado)
    print("¡Plantilla creada con éxito en:", ruta_guardado)

if __name__ == "__main__":
    crear_plantilla_excel()